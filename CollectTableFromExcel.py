import argparse
import os
import json
import sys
import io
import openpyxl
from openpyxl.utils import range_boundaries

def log(message, verbose):
    if verbose:
        sys.stderr.write(f"[LOG] {message}\n")

def resolve_range(wb, ws, range_name):
    """
    Tìm kiếm tọa độ từ RangeName. 
    Hỗ trợ cả Named Range (Defined Names) và Excel Table.
    """
    # 1. Thử tìm trong Named Ranges (Name Manager) - Toàn bộ Workbook
    if range_name in wb.defined_names:
        defn = wb.defined_names[range_name]
        dests = list(defn.destinations)
        if dests:
            sheet_name, coord = dests[0]
            # Trả về tọa độ, loại bỏ dấu $
            return coord.replace('$', '')

    # 2. Thử tìm trong danh sách Table (Insert > Table) - Chỉ trong Sheet hiện tại
    if range_name in ws.tables:
        return ws.tables[range_name].ref

    return None

def extract_table_data(wb, sheet_spec, verbose):
    sheet_name = sheet_spec.get("name")
    if sheet_name not in wb.sheetnames:
        log(f"Không tìm thấy sheet: {sheet_name}", verbose)
        return []

    ws = wb[sheet_name]
    table_spec = sheet_spec.get("table", {})
    mapping_fields = table_spec.get("MappingFields", [])
    
    # Xác định vùng dữ liệu
    range_str = None
    r_name = table_spec.get("RangeName")
    
    if r_name:
        range_str = resolve_range(wb, ws, r_name)
        if range_str:
            log(f"Đã tìm thấy vùng cho '{r_name}': {range_str}", verbose)
        else:
            log(f"Không tìm thấy Named Range hoặc Table tên '{r_name}'", verbose)
    
    # Nếu không tìm thấy theo tên, dùng Range cứng trong JSON
    if not range_str:
        range_str = table_spec.get("Range")
    
    if not range_str or range_str == r_name: # Tránh trường hợp range_str vẫn bị gán ngược lại tên
        log(f"LỖI: Không có tọa độ hợp lệ để đọc dữ liệu.", verbose)
        return []

    table_data = []
    try:
        # ws[range_str] bây giờ chắc chắn là dạng "A1:C10"
        for row in ws[range_str]:
            row_dict = {}
            for idx, cell in enumerate(row):
                if idx < len(mapping_fields):
                    field_name = mapping_fields[idx]
                    row_dict[field_name] = cell.value
            table_data.append(row_dict)
    except Exception as e:
        log(f"Lỗi khi đọc vùng {range_str}: {e}", verbose)

    return table_data

def main():
    parser = argparse.ArgumentParser(description="Trích xuất dữ liệu bảng từ Excel sang JSON.")
    parser.add_argument("-d", "--document", required=True, help="Đường dẫn file .xlsx hoặc .xls")
    parser.add_argument("-j", "--json", help="File JSON đặc tả cấu trúc")
    parser.add_argument("-o", "--output", help="File đầu ra .json")
    parser.add_argument("-v", "--verbose", action="store_true", help="Hiển thị chi tiết quá trình")

    args = parser.parse_args()

    # 1. Kiểm tra file Excel
    if not os.path.exists(args.document):
        sys.stderr.write(f"Lỗi: File '{args.document}' không tồn tại.\n")
        sys.exit(1)

    # 2. Xác định file JSON cấu trúc
    json_path = args.json if args.json else os.path.splitext(args.document)[0] + ".json"
    if not os.path.exists(json_path):
        sys.stderr.write(f"Lỗi: File cấu trúc '{json_path}' không tồn tại.\n")
        sys.exit(1)

    # 3. Đọc file đặc tả
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
    except Exception as e:
        sys.stderr.write(f"Lỗi đọc file JSON: {e}\n")
        sys.exit(1)

    # 4. Mở Excel và xử lý
    results = {}
    try:
        # data_only=True để lấy giá trị cuối cùng thay vì công thức
        wb = openpyxl.load_workbook(args.document, data_only=True)
        
        for sheet_cfg in config.get("sheets", []):
            s_name = sheet_cfg.get("name")
            log(f"Đang xử lý sheet: {s_name}", args.verbose)
            results[s_name] = extract_table_data(wb, sheet_cfg, args.verbose)

        # 5. Xuất kết quả
        output_json = json.dumps(results, indent=4, ensure_ascii=False)
        
        if args.output:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(output_json)
            log(f"Đã ghi kết quả ra file: {args.output}", args.verbose)
        else:
            # Ép kiểu stdout sang utf-8 để in tiếng Việt chuẩn
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
            sys.stdout.write(output_json + "\n")

    except Exception as e:
        sys.stderr.write(f"Lỗi hệ thống: {e}\n")
        sys.exit(1)

if __name__ == "__main__":
    main()