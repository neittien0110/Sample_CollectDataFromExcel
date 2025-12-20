import argparse
import os
import json
import sys
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
from typing import Dict, Any, List

# --- CÁC HÀM TIỆN ÍCH ---

def get_file_components(file_path: str) -> (str, str, str):
    """Phân tách đường dẫn thành thư mục, tên gốc và đuôi mở rộng."""
    directory = os.path.dirname(os.path.abspath(file_path))
    full_name = os.path.basename(file_path)
    name_without_ext, ext = os.path.splitext(full_name)
    return directory, name_without_ext, ext

def map_named_ranges(wb) -> Dict[str, str]:
    """Tạo ánh xạ (map) từ Tên đã đặt (Named Range) sang tọa độ A1 của ô đơn."""
    named_ranges_map = {}
    
    for name, defined_name in wb.defined_names.items(): 
        # Lấy destinations (là một generator)
        destinations_gen = defined_name.destinations
        
        # CHUYỂN GENERATOR THÀNH LIST để có thể kiểm tra len()
        destinations = list(destinations_gen)
        
        if destinations and len(destinations) == 1:
            sheet, address = destinations[0]
            
            # Chỉ lấy các tên trỏ đến 1 ô đơn (single cell)
            if ':' not in address: 
                named_ranges_map[name] = address
                
    return named_ranges_map

def get_cell_address(location_name: str, named_ranges_map: Dict[str, str], verbose: bool) -> str:
    """Chuyển đổi vị trí (A5 hoặc Named Range) thành tọa độ A1 hợp lệ."""
    
    # 1. KIỂM TRA NAMED RANGE (Ưu tiên)
    if location_name in named_ranges_map:
        if verbose: print(f"    - Map: Đã ánh xạ Named Range '{location_name}' sang tọa độ {named_ranges_map[location_name]}")
        return named_ranges_map[location_name]

    # 2. KIỂM TRA CÚ PHÁP A1
    try:
        # Nếu không phải Named Range, thử xem nó có phải là tọa độ A1 hợp lệ không
        coordinate_from_string(location_name)
        return location_name
    except ValueError:
        # 3. NẾU CẢ HAI ĐỀU SAI
        raise ValueError(f"Vị trí không hợp lệ: '{location_name}' (Không phải A1 và không phải Named Range)")

def map_image_names_and_locations(excel_path: str, wb, verbose: bool) -> Dict[str, Dict[str, Any]]:
    """
    Sử dụng XML Parsing để ánh xạ Tên Đối tượng (Object Name) tới vị trí neo đậu,
    linh hoạt tìm kiếm tên qua tất cả các cấu trúc (pic, shape, contentPart).
    """
    image_map = {}
    sheet_id_map = {ws.title: str(i + 1) for i, ws in enumerate(wb.worksheets)}
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            
            # Khai báo các Namespace cần thiết
            xdr_ns = '{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetdrawing}'
            
            # 1. Ánh xạ Drawing ID và file XML
            # (Phần này giữ nguyên, đã ổn định)
            rels_root = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
            drawing_rels = {}
            for rel in rels_root.findall(".//{*}Relationship"):
                 if rel.get('Type') and rel.get('Type').endswith('drawing'):
                    drawing_rels[rel.get('Id')] = rel.get('Target').replace('../', 'xl/')

            # 2. Xử lý từng sheet để tìm Drawing File
            for sheet_name, sheet_id in sheet_id_map.items():
                sheet_rel_file = f'xl/worksheets/_rels/sheet{sheet_id}.xml.rels'
                if sheet_rel_file not in z.namelist(): continue
                
                sheet_rels_root = ET.fromstring(z.read(sheet_rel_file))
                
                drawing_id = None
                for rel in sheet_rels_root.findall(".//{*}Relationship"):
                    if rel.get('Type') and rel.get('Type').endswith('drawing'):
                        drawing_id = rel.get('Id')
                        break
                
                if not drawing_id or drawing_id not in drawing_rels: continue
                    
                drawing_file = drawing_rels[drawing_id]
                
                # 3. Phân tích file Drawing để lấy Tên Đối tượng và Vị trí
                if drawing_file in z.namelist():
                    drawing_root = ET.fromstring(z.read(drawing_file))
                    
                    for anchor in drawing_root.findall(f'{xdr_ns}twoCellAnchor'):
                        
                        # --- Trích xuất Vị trí Neo đậu ---
                        from_tag = anchor.find(f'{xdr_ns}from')
                        
                        # Tọa độ A1 (Col: 6 -> G, Row: 85 -> 86. Col: 7 -> H, Row: 74 -> 75. Col: 2 -> C, Row: 72 -> 73)
                        # Lưu ý: XML lưu chỉ mục 0, Excel hiển thị chỉ mục 1, nên ta +1 cho Row.
                        try:
                            col_from = from_tag.find(f'{xdr_ns}col').text
                            row_from = from_tag.find(f'{xdr_ns}row').text
                            from_cell_coords = f'{coordinate_from_string(col_from)[0]}{int(row_from) + 1}'
                        except AttributeError:
                             if verbose: print(f"    - [XML] Lỗi đọc tọa độ neo đậu.")
                             continue
                        
                        # --- Trích xuất Tên Đối tượng (Object Name) - Sử dụng tìm kiếm toàn cục ---
                        image_name = None
                        
                        # Tìm tất cả các thẻ *cNvPr* bên dưới twoCellAnchor. 
                        # Chúng ta cần tìm trong tất cả các namespace có thể (xdr, xdr14)
                        
                        # ET.findall(tag) chỉ tìm kiếm các thẻ con trực tiếp, ET.iter(tag) tìm kiếm đệ quy.
                        # Do các namespace khác nhau, ta phải dùng cách chung nhất:
                        for c_nv_pr in anchor.iter():
                            if c_nv_pr.tag.endswith('cNvPr'):
                                # Tên Đối tượng luôn nằm trong thuộc tính 'name'
                                name_attr = c_nv_pr.get('name')
                                if name_attr:
                                    image_name = name_attr
                                    break
                                    
                        # Lưu ánh xạ
                        if image_name:
                            image_map[image_name] = {
                                'sheet': sheet_name,
                                'location': from_cell_coords,
                            }
                            if verbose: print(f"  [XML] Ánh xạ ảnh: {image_name} (Sheet: {sheet_name}, Vị trí: {from_cell_coords})")
                            
        return image_map

    except Exception as e:
        if verbose: print(f"  [XML] Lỗi khi phân tích XML để lấy tên ảnh: {e}")
        return image_map
    
# --- HÀM TRÍCH XUẤT CHÍNH ---

def extract_excel_data(excel_path: str, spec_data: Dict[str, Any], output_media_dir: str, file_prefix: str, verbose: bool) -> Dict[str, Any]:
    
    final_results = {}
    image_counter = 0
    
    try:
        wb = load_workbook(excel_path, data_only=True) 
    except Exception as e:
        print(f"LỖI: Không thể mở file Excel {excel_path} với openpyxl: {e}")
        return {}
    
    # 1. Ánh xạ các Named Range và Tên Đối tượng
    named_ranges_map = map_named_ranges(wb)
    
    image_spec_check = any('images' in s for s in spec_data.get('sheets', []))
    named_image_map = map_image_names_and_locations(excel_path, wb, verbose) if image_spec_check else {}

    # Xử lý từng Sheet trong đặc tả
    for sheet_spec in spec_data.get('sheets', []):
        sheet_name = sheet_spec.get('name')
        
        if not sheet_name or sheet_name not in wb.sheetnames:
            if verbose: print(f"\n  [VERBOSE] Bỏ qua: Sheet '{sheet_name}' không tồn tại.")
            continue
        
        ws = wb[sheet_name]
        if verbose: print(f"\n  [VERBOSE] Đang xử lý Sheet: {sheet_name}")

        # --- 2. Trích xuất Cell Data (A1 hoặc Named Range) ---
        for cell_item in sheet_spec.get('cells', []):
            spec_name = cell_item['name']
            location = cell_item['location']
            
            try:
                a1_location = get_cell_address(location, named_ranges_map, verbose)
                cell_value = ws[a1_location].value
                final_results[spec_name] = cell_value
                if verbose: print(f"    - CELL: Đã trích xuất '{spec_name}' tại {location} ({a1_location}): {cell_value}")
                
            except Exception as e:
                final_results[spec_name] = None
                if verbose: print(f"    - LỖI CELL: Không thể đọc ô {location}. Lỗi: {e}")
        
        # --- 3. Trích xuất Image Objects (theo Object Name) ---
        for image_item in sheet_spec.get('images', []):
            spec_name = image_item['name'] 
            object_name = image_item['location'] # Tên đối tượng (ví dụ: Graphic 1)

            is_found = False
            target_location = None
            
            # Tìm kiếm vị trí neo đậu A1 từ Tên Đối tượng
            if object_name in named_image_map and named_image_map[object_name]['sheet'] == sheet_name:
                target_location = named_image_map[object_name]['location']
            else:
                if verbose: print(f"    - LỖI IMAGE: Không tìm thấy Tên Đối tượng '{object_name}' trong XML hoặc không nằm trên sheet này.")
                final_results[spec_name] = None
                continue

            # Lặp qua các đối tượng Image của openpyxl để trích xuất
            for img in getattr(ws, '_images', []):
                anchor_cell = img.anchor.from_cell.col_letter + str(img.anchor.from_cell.row)
                
                # So sánh vị trí neo đậu với vị trí của Object Name
                if anchor_cell == target_location:
                    image_counter += 1
                    is_found = True
                    
                    ext = img.path.split('.')[-1].lower() if img.path else "png" 
                    output_filename = f"{file_prefix}_{image_counter:02d}.{ext}"
                    output_path = os.path.join(output_media_dir, output_filename)
                    
                    with open(output_path, 'wb') as f:
                        f.write(img.ref.file.read())
                    
                    final_results[spec_name] = output_path
                    if verbose: print(f"    - IMAGE: Đã lưu ảnh '{spec_name}' (Tên Đối tượng: {object_name}) -> {output_filename}")
                    
                    break
            
            if verbose and not is_found:
                print(f"    - IMAGE: Tên Đối tượng '{object_name}' đã được ánh xạ ({target_location}), nhưng không tìm thấy đối tượng ảnh openpyxl tương ứng.")
                final_results[spec_name] = None
                
    return final_results

# --- HÀM CHÍNH ---

def main():
    parser = argparse.ArgumentParser(description="Trích xuất dữ liệu Excel.")
    parser.add_argument('--document', '-d', type=str, required=True, help='Tên file Excel nguồn dữ liệu (ví dụ: abc.xlsx).')
    
    # Tham số --json (-j) TÙY CHỌN, có MẶC ĐỊNH
    parser.add_argument('--json', '-j', type=str, default=None, 
                        help='Tên file JSON đặc tả cấu trúc. Mặc định: [tên document].json.')

    parser.add_argument('--output', '-o', type=str, default=None, help='File JSON đầu ra. Nếu trống sẽ in ra màn hình.')
   
    # Tham số --verbose (-v) CỜ BOOLEAN
    parser.add_argument('--verbose', '-v', action='store_true', default=False, 
                        help='Giải thích chi tiết từng bước hoạt động.')
    
    args = parser.parse_args()
    
    # 1. Kiểm tra file đầu vào
    if not os.path.exists(args.document):
        sys.exit(1)

    doc_dir, doc_name_no_ext, _ = get_file_components(args.document)
    json_path = args.json if args.json else os.path.join(doc_dir, f"{doc_name_no_ext}.json")
    
    if not os.path.exists(json_path):
        sys.exit(1)
        
    output_media_dir = os.path.join(doc_dir, f"{doc_name_no_ext}_media")
    os.makedirs(output_media_dir, exist_ok=True) 

    # --- 2. Đọc Đặc tả JSON và Trích xuất ---
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            spec_data = json.load(f)
    except Exception as e:
        print(f"LỖI: Không thể đọc hoặc phân tích file JSON đặc tả: {e}")
        sys.exit(1)
        
    final_results = extract_excel_data(args.document, spec_data, output_media_dir, doc_name_no_ext, args.verbose)
    
    # 3. Xuất kết quả
    json_output = json.dumps(final_results, indent=4, ensure_ascii=False)
    
    if args.output:
        try:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(json_output)
            if args.verbose:
                print(f"THÀNH CÔNG: Đã lưu vào {args.output}")
        except Exception as e:
            sys.stderr.write(f"LỖI: {str(e)}\n")
    else:
        # CHỈ in nội dung JSON, không in gì khác
        sys.stdout.write(json_output + "\n")

if __name__ == "__main__":
    main()